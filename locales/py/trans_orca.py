#python
from openpyxl import load_workbook
from xml.etree.ElementTree import Element, SubElement, tostring
import xml.dom.minidom
import os

def parse_excel_to_ts(excel_file, sheet_name, ts_file):
    # 读取Excel文件
    wb = load_workbook(excel_file)
    sheet = wb[sheet_name]

    # 创建TS文件的根元素
    root = Element('TS')
    root.set('version', '2.1')
    context = SubElement(root, 'context')
    SubElement(context, 'name').text = "cxTranslation"

    # 遍历excel内容
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row[0] and row[2]:
            message = SubElement(context, 'message')
            source = SubElement(message, 'source')
            source.text = row[0]
            translation = SubElement(message, 'translation')
            translation.text = row[2]
    
    # 将生成的XML转换为字符串
    xml_str = tostring(root, encoding='utf-8')
    dom = xml.dom.minidom.parseString(xml_str)
    pretty_xml_as_string = dom.toprettyxml()

    # 写入文件到ts_file
    with open(ts_file, 'w', encoding='utf-8') as f:
        f.write(pretty_xml_as_string)

# 示例使用
sheet_name = 'Sheet1'

# 指定文件夹路径
excel_folder_path = '../excel/orca/'
ts_folder_path = '../ts_new/orca/'

# 获取文件夹下所有文件
files = os.listdir(excel_folder_path)

if not os.path.exists(ts_folder_path):
    os.makedirs(ts_folder_path)
    
# 输出每个文件的前缀
suffix = "_orca_new"

print(f"\n--------开始生成ts_orca--------")

for file in files:
    if os.path.isfile(os.path.join(excel_folder_path, file)):
        prefix = os.path.splitext(file)[0]
        parse_excel_to_ts(f"{excel_folder_path}{prefix}.xlsx", sheet_name, f"{ts_folder_path}{prefix}{suffix}.ts")
        print(f"生成{ts_folder_path}{prefix}.ts")
        
print(f"--------结束生成ts_orca--------")